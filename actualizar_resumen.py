import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys

def actualizar_resumen(file_path: str, csv_path: str = None):
    """
    Procesa un archivo Excel con varias hojas de notificaciones y genera un RESUMEN.
    Si se pasa un CSV, lo agrega como nueva hoja antes de actualizar el resumen.
    """
    # --- 1. Leer Excel ---
    wb = load_workbook(file_path)

    # --- 2. Si hay CSV, lo agregamos como nueva hoja ---
    df_csv = None
    fecha = None
    if csv_path:
        df_csv = pd.read_csv(csv_path)
        df_csv.columns = df_csv.columns.str.strip()

        # Tomamos la fecha de la primera fila como nombre de hoja
        fecha = str(df_csv["FECHA"].iloc[0])
        if fecha in wb.sheetnames:
            # Si ya existe, eliminamos para actualizarla
            wb.remove(wb[fecha])

        ws_nueva = wb.create_sheet(fecha)
        for r in dataframe_to_rows(df_csv, index=False, header=True):
            ws_nueva.append(r)

    # --- 3. Eliminar hoja RESUMEN si existe ---
    if "RESUMEN" in wb.sheetnames:
        wb.remove(wb["RESUMEN"])
    ws_resumen = wb.create_sheet("RESUMEN")

    # --- 4. Procesar todas las hojas y armar resumen ---
    resumen_data = []
    for sheet in wb.sheetnames:
        if sheet.upper() == "RESUMEN":
            continue

        if csv_path and sheet == fecha:
            df = df_csv
        else:
            df = pd.read_excel(file_path, sheet_name=sheet)
        df.columns = df.columns.str.strip()

        mensaje_enviado = df[((df['TELEFONO'].notna()) & (df['TELEFONO_2'].isna())) |
                             ((df['TELEFONO'].isna()) & (df['TELEFONO_2'].notna()))].shape[0]

        mensaje_no_enviado = df[(df['TELEFONO'].isna()) & (df['TELEFONO_2'].isna())].shape[0]

        mensaje_dos_numeros = df[(df['TELEFONO'].notna()) & (df['TELEFONO_2'].notna())].shape[0]

        resumen_data.append([sheet, mensaje_enviado, mensaje_no_enviado, mensaje_dos_numeros])

    # --- 5. Escribir RESUMEN en Excel ---
    headers = ["Fecha", "Mensaje enviado", "Mensaje no enviado", "Se envió a los dos números"]
    ws_resumen.append(headers)
    for row in resumen_data:
        ws_resumen.append(row)

    # Crear tabla formateada
    end_row = ws_resumen.max_row
    table = Table(displayName="ResumenTable", ref=f"A1:D{end_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws_resumen.add_table(table)

    # --- 6. Guardar archivo final (sobrescribiendo el original) ---
    wb.save(file_path)
    print(f"Archivo actualizado: {file_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python actualizar_resumen.py <archivo_excel> [archivo_csv]")
    else:
        file_path = sys.argv[1]
        csv_path = sys.argv[2] if len(sys.argv) > 2 else None
        actualizar_resumen(file_path, csv_path)
